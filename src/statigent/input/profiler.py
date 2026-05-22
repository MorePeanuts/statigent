"""Input file discovery and dataset-shape-specific profiling.

The InputProfiler is the first stage of the data science pipeline:
1. Discover files from paths, directories, and zip archives.
2. Classify the discovered dataset shape.
3. Dispatch to a shape-specific profiler.
4. Return a DatasetProfile whose compact_summary is tailored to the shape.

Security: zip extraction validates that every member stays within the
target directory (path-traversal protection) and enforces total
uncompressed size limits.
"""

import csv
import zipfile
from collections import Counter
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from os.path import commonpath
from pathlib import Path
from shutil import copyfileobj, rmtree
from typing import Protocol

import pandas as pd
from openpyxl import load_workbook  # type: ignore[import-untyped]
from pandas.api.types import is_datetime64_any_dtype, is_numeric_dtype
from pandas.errors import ParserError
from PIL import Image, UnidentifiedImageError

from statigent.errors import StatigentInputError
from statigent.schemas import (
    DatasetKind,
    DatasetProfile,
    ImageCollectionProfile,
    InputFileInfo,
    SpreadsheetSheetProfile,
    SpreadsheetWorkbookProfile,
    TableProfile,
    TableRole,
)

TABULAR_SUFFIXES = frozenset({".csv", ".tsv", ".xlsx", ".xls", ".parquet"})
EXCEL_WORKBOOK_SUFFIXES = frozenset({".xlsx"})
IMAGE_SUFFIXES = frozenset({".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"})
ZIP_COPY_CHUNK_BYTES = 1024 * 1024


class _TableProfilingError(Exception):
    """Expected failure while deriving pandas profile statistics."""


@dataclass(frozen=True)
class _DiscoveredFile:
    path: Path
    relative_path: str


@dataclass(frozen=True)
class _DiscoveryResult:
    files: list[_DiscoveredFile]
    input_file_infos: list[InputFileInfo]
    warnings: list[str]


@dataclass(frozen=True)
class _ProfileContext:
    files: list[_DiscoveredFile]
    input_file_infos: list[InputFileInfo]
    warnings: list[str]


class _DatasetProfiler(Protocol):
    def profile(self, context: _ProfileContext) -> DatasetProfile: ...


class _WorksheetLike(Protocol):
    title: str
    max_row: int
    max_column: int

    def iter_rows(
        self,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
        values_only: bool = False,
    ) -> Iterable[tuple[object, ...]]: ...


class _EmptyDatasetProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.EMPTY,
            files=list(context.input_file_infos),
            tables=[],
            warnings=list(context.warnings),
        )


class _SingleTableProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        files, tables, warnings = self.owner._profile_tabular_files(
            context,
            role_detector=self.owner._table_role,
        )
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.SINGLE_TABLE,
            files=files,
            tables=tables,
            warnings=warnings,
        )


class _MultiTableProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        files, tables, warnings = self.owner._profile_tabular_files(
            context,
            role_detector=self.owner._table_role,
        )
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.MULTI_TABLE,
            files=files,
            tables=tables,
            warnings=warnings,
        )


class _ModelingSplitTableProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        files, tables, warnings = self.owner._profile_tabular_files(
            context,
            role_detector=self.owner._table_role,
        )
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.MODELING_SPLIT_TABLES,
            files=files,
            tables=tables,
            warnings=warnings,
        )


class _ImageCollectionProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        files = self.owner._input_file_infos(context)
        collection, warnings = self.owner._profile_image_collection(context)
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.IMAGE_COLLECTION,
            files=files,
            tables=[],
            image_collections=[collection] if collection is not None else [],
            warnings=[*context.warnings, *warnings],
        )


class _SpreadsheetWorkbookProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        files = self.owner._input_file_infos(context)
        workbooks, warnings = self.owner._profile_spreadsheet_workbooks(context)
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.SPREADSHEET_WORKBOOK,
            files=files,
            tables=[],
            spreadsheet_workbooks=workbooks,
            warnings=[*context.warnings, *warnings],
        )


class _FallbackProfiler:
    def __init__(self, owner: "InputProfiler") -> None:
        self.owner = owner

    def profile(self, context: _ProfileContext) -> DatasetProfile:
        files, tables, warnings = self.owner._profile_tabular_files(
            context,
            role_detector=self.owner._table_role,
        )
        return DatasetProfile(
            root=self.owner.work_dir,
            kind=DatasetKind.MIXED,
            files=files,
            tables=tables,
            warnings=warnings,
        )


class InputProfiler:
    """Discover and profile input files for a data science task.

    Handles single files, directories (recursive), and zip archives.
    A single profiling failure does not abort the run — the file is skipped
    and the warning is recorded in the DatasetProfile.
    """

    def __init__(
        self,
        work_dir: Path,
        max_files: int = 200,
        max_file_bytes: int = 250_000_000,
        sample_rows: int = 5,
    ) -> None:
        self.work_dir = work_dir
        self.max_files = max_files
        self.max_file_bytes = max_file_bytes
        self.sample_rows = sample_rows

    def profile_paths(self, paths: list[Path] | None) -> DatasetProfile:
        """Profile all discovered files and return a DatasetProfile.

        Accepts None (no inputs), single files, directories, or zip archives.
        Each tabular file is profiled independently — failures are recorded
        as warnings rather than crashing the run.
        """
        warnings: list[str] = []
        discovery = self._discover_paths(paths, warnings)
        context = _ProfileContext(
            files=discovery.files,
            input_file_infos=discovery.input_file_infos,
            warnings=[*warnings, *discovery.warnings],
        )
        profiler = self._select_profiler(context)
        return profiler.profile(context)

    def _discover_paths(
        self, paths: list[Path] | None, warnings: list[str]
    ) -> _DiscoveryResult:
        if paths is None:
            return _DiscoveryResult(files=[], input_file_infos=[], warnings=[])

        discovered: list[_DiscoveredFile] = []
        input_file_infos: list[InputFileInfo] = []
        discovery_warnings: list[str] = []
        for raw_path in paths:
            path = raw_path.expanduser()
            if not path.exists():
                msg = f"Input path does not exist: {path}"
                raise StatigentInputError(msg)
            if self._is_ignored_input_file(path):
                continue

            if path.is_dir():
                self._add_discovered(
                    self._scan_directory(path),
                    discovered,
                    warnings,
                )
            elif path.suffix.lower() == ".zip":
                zip_result = self._extract_zip(path)
                input_file_infos.extend(zip_result.input_file_infos)
                discovery_warnings.extend(zip_result.warnings)
                self._add_discovered(
                    zip_result.files,
                    discovered,
                    warnings,
                )
            elif path.is_file():
                self._add_discovered(
                    [_DiscoveredFile(path=path, relative_path=path.name)],
                    discovered,
                    warnings,
                )

        return _DiscoveryResult(
            files=discovered,
            input_file_infos=input_file_infos,
            warnings=discovery_warnings,
        )

    def _add_discovered(
        self,
        items: Iterable[_DiscoveredFile],
        discovered: list[_DiscoveredFile],
        warnings: list[str],
    ) -> None:
        for item in items:
            if len(discovered) >= self.max_files:
                warning = f"Stopped input discovery at max_files={self.max_files}."
                warnings.append(warning)
                return
            discovered.append(item)

    def _is_ignored_input_file(self, path: Path) -> bool:
        return path.name.startswith("~$")

    def _scan_directory(self, directory: Path) -> list[_DiscoveredFile]:
        items: list[_DiscoveredFile] = []
        work_dir = self.work_dir.resolve()
        for path in sorted(directory.rglob("*")):
            if not path.is_file():
                continue
            if self._is_ignored_input_file(path):
                continue
            if self._is_relative_to(path.resolve(), work_dir):
                continue
            items.append(
                _DiscoveredFile(
                    path=path,
                    relative_path=path.relative_to(directory).as_posix(),
                )
            )
        return items

    def _extract_zip(self, archive: Path) -> _DiscoveryResult:
        inputs_root = (self.work_dir / "inputs").resolve()
        target_dir = self.work_dir / "inputs" / archive.stem
        target_root = target_dir.resolve()
        self._prepare_zip_target_dir(target_dir, inputs_root)
        skipped_file_infos: list[InputFileInfo] = []
        discovery_warnings: list[str] = []

        try:
            with zipfile.ZipFile(archive) as zf:
                members = zf.infolist()
                if len(members) > self.max_files:
                    discovery_warnings.append(
                        f"Only the first {self.max_files} out of {len(members)} "
                        "zip members were scanned."
                    )
                total_uncompressed_size = 0
                for member in members[: self.max_files]:
                    member_relative_path = f"{archive.name}/{member.filename}"
                    destination = (target_dir / member.filename).resolve()
                    if not self._is_relative_to(destination, target_root):
                        msg = f"Zip archive contains unsafe path: {member.filename}"
                        raise StatigentInputError(msg)
                    if member.file_size > self.max_file_bytes:
                        suffix = Path(member.filename).suffix.lower()
                        skipped_file_infos.append(
                            InputFileInfo(
                                path=target_dir / member.filename,
                                relative_path=member_relative_path,
                                suffix=suffix,
                                size_bytes=member.file_size,
                                is_tabular=suffix in TABULAR_SUFFIXES,
                            )
                        )
                        discovery_warnings.append(
                            f"Skipped {member_relative_path}: file exceeds "
                            f"max_file_bytes={self.max_file_bytes}."
                        )
                        continue
                    total_uncompressed_size += member.file_size
                    max_uncompressed_size = self.max_files * self.max_file_bytes
                    if total_uncompressed_size > max_uncompressed_size:
                        msg = (
                            "Zip archive uncompressed size exceeds limit: "
                            f"{total_uncompressed_size} > {max_uncompressed_size}."
                        )
                        raise StatigentInputError(msg)

                    if member.is_dir():
                        destination.mkdir(parents=True, exist_ok=True)
                        continue
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    with zf.open(member) as source, destination.open("wb") as output:
                        copyfileobj(source, output, length=ZIP_COPY_CHUNK_BYTES)
        except zipfile.BadZipFile as err:
            msg = f"Invalid zip archive: {archive}"
            raise StatigentInputError(msg) from err

        return _DiscoveryResult(
            files=[
                _DiscoveredFile(
                    path=path,
                    relative_path=(
                        f"{archive.name}/{path.relative_to(target_dir).as_posix()}"
                    ),
                )
                for path in sorted(target_dir.rglob("*"))
                if path.is_file()
            ],
            input_file_infos=skipped_file_infos,
            warnings=discovery_warnings,
        )

    def _prepare_zip_target_dir(self, target_dir: Path, inputs_root: Path) -> None:
        target_root = target_dir.resolve()
        if not self._is_relative_to(target_root, inputs_root):
            msg = f"Unsafe zip extraction target: {target_dir}"
            raise StatigentInputError(msg)
        if target_dir.exists():
            if target_dir.is_dir():
                rmtree(target_dir)
            else:
                target_dir.unlink()

    def _select_profiler(self, context: _ProfileContext) -> _DatasetProfiler:
        if not context.files and not context.input_file_infos:
            return _EmptyDatasetProfiler(self)

        tabular_files = [
            item
            for item in context.files
            if item.path.suffix.lower() in TABULAR_SUFFIXES
        ]
        image_files = [
            item for item in context.files if item.path.suffix.lower() in IMAGE_SUFFIXES
        ]

        if image_files and len(image_files) == len(context.files):
            return _ImageCollectionProfiler(self)
        if self._has_spreadsheet_grid_workbook(context.files):
            return _SpreadsheetWorkbookProfiler(self)
        if not tabular_files:
            return _FallbackProfiler(self)
        if len(tabular_files) != len(context.files):
            return _FallbackProfiler(self)

        roles = {self._table_role(item.relative_path) for item in tabular_files}
        if TableRole.TRAIN in roles and TableRole.TEST in roles:
            return _ModelingSplitTableProfiler(self)
        if TableRole.TRAIN in roles and TableRole.VALIDATION in roles:
            return _ModelingSplitTableProfiler(self)

        logical_table_count = self._logical_table_count(tabular_files)
        if logical_table_count == 1:
            return _SingleTableProfiler(self)
        return _MultiTableProfiler(self)

    def _input_file_infos(self, context: _ProfileContext) -> list[InputFileInfo]:
        files = list(context.input_file_infos)
        for item in context.files:
            suffix = item.path.suffix.lower()
            files.append(
                InputFileInfo(
                    path=item.path,
                    relative_path=item.relative_path,
                    suffix=suffix,
                    size_bytes=item.path.stat().st_size,
                    is_tabular=suffix in TABULAR_SUFFIXES,
                )
            )
        return files

    def _profile_tabular_files(
        self,
        context: _ProfileContext,
        *,
        role_detector: Callable[[str], TableRole],
    ) -> tuple[list[InputFileInfo], list[TableProfile], list[str]]:
        files = self._input_file_infos(context)
        tables: list[TableProfile] = []
        warnings = list(context.warnings)

        for item in context.files:
            suffix = item.path.suffix.lower()
            if suffix not in TABULAR_SUFFIXES:
                continue
            size_bytes = item.path.stat().st_size
            if size_bytes > self.max_file_bytes:
                warning = (
                    f"Skipped table profiling for {item.relative_path}: "
                    f"{size_bytes} bytes exceeds max_file_bytes={self.max_file_bytes}."
                )
                warnings.append(warning)
                continue

            try:
                tables.extend(self._profile_logical_tables(item, role_detector))
            except (
                OSError,
                ValueError,
                ImportError,
                ParserError,
                _TableProfilingError,
                zipfile.BadZipFile,
            ) as err:
                warning = f"Failed to profile {item.relative_path}: {err}"
                warnings.append(warning)
                continue

        return files, tables, warnings

    def _profile_logical_tables(
        self,
        item: _DiscoveredFile,
        role_detector: Callable[[str], TableRole],
    ) -> list[TableProfile]:
        suffix = item.path.suffix.lower()
        if suffix in EXCEL_WORKBOOK_SUFFIXES:
            workbook = pd.ExcelFile(item.path)
            return [
                self._profile_table(
                    item,
                    frame=pd.read_excel(workbook, sheet_name=sheet_name),
                    relative_path=f"{item.relative_path}::{sheet_name}",
                    source_label=f"{item.relative_path}::{sheet_name}",
                    role=role_detector(f"{item.relative_path}::{sheet_name}"),
                )
                for sheet_name in workbook.sheet_names
            ]

        return [
            self._profile_table(
                item,
                frame=self._read_table(item.path),
                relative_path=item.relative_path,
                source_label=item.relative_path,
                role=role_detector(item.relative_path),
            )
        ]

    def _logical_table_count(self, tabular_files: list[_DiscoveredFile]) -> int:
        count = 0
        for item in tabular_files:
            if item.path.suffix.lower() not in EXCEL_WORKBOOK_SUFFIXES:
                count += 1
                continue
            try:
                count += len(pd.ExcelFile(item.path).sheet_names)
            except (OSError, ValueError, ImportError, zipfile.BadZipFile):
                count += 1
        return count

    def _has_spreadsheet_grid_workbook(self, files: list[_DiscoveredFile]) -> bool:
        workbook_files = [
            item
            for item in files
            if item.path.suffix.lower() in EXCEL_WORKBOOK_SUFFIXES
        ]
        if not workbook_files or len(workbook_files) != len(files):
            return False
        return any(self._workbook_looks_like_grid(item) for item in workbook_files)

    def _workbook_looks_like_grid(self, item: _DiscoveredFile) -> bool:
        try:
            workbook = load_workbook(item.path, read_only=True, data_only=True)
        except (OSError, ValueError, zipfile.BadZipFile):
            return False
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    non_empty = [
                        value
                        for value in row
                        if value is not None and str(value).strip() != ""
                    ]
                    if not non_empty:
                        continue
                    return len(non_empty) <= 1
        finally:
            workbook.close()
        return False

    def _profile_spreadsheet_workbooks(
        self,
        context: _ProfileContext,
    ) -> tuple[list[SpreadsheetWorkbookProfile], list[str]]:
        workbooks: list[SpreadsheetWorkbookProfile] = []
        warnings: list[str] = []
        for item in context.files:
            if item.path.suffix.lower() not in EXCEL_WORKBOOK_SUFFIXES:
                continue
            try:
                workbooks.append(self._profile_spreadsheet_workbook(item))
            except (OSError, ValueError, zipfile.BadZipFile) as err:
                warnings.append(
                    "Failed to profile spreadsheet workbook "
                    f"{item.relative_path}: {err}"
                )
        return workbooks, warnings

    def _profile_spreadsheet_workbook(
        self,
        item: _DiscoveredFile,
    ) -> SpreadsheetWorkbookProfile:
        values_workbook = load_workbook(item.path, read_only=True, data_only=True)
        formulas_workbook = load_workbook(item.path, read_only=True, data_only=False)
        try:
            formula_sheets = {
                worksheet.title: worksheet for worksheet in formulas_workbook.worksheets
            }
            sheets = [
                self._profile_spreadsheet_sheet(
                    values_worksheet=worksheet,
                    formula_worksheet=formula_sheets[worksheet.title],
                )
                for worksheet in values_workbook.worksheets
            ]
        finally:
            values_workbook.close()
            formulas_workbook.close()
        return SpreadsheetWorkbookProfile(
            path=item.path,
            relative_path=item.relative_path,
            sheets=sheets,
        )

    def _profile_spreadsheet_sheet(
        self,
        *,
        values_worksheet: _WorksheetLike,
        formula_worksheet: _WorksheetLike,
    ) -> SpreadsheetSheetProfile:
        row_count = int(values_worksheet.max_row or 0)
        column_count = int(values_worksheet.max_column or 0)
        non_empty_cells = self._count_non_empty_cells(values_worksheet)
        formula_cells = self._count_formula_cells(formula_worksheet)
        return SpreadsheetSheetProfile(
            name=values_worksheet.title,
            rows=row_count,
            columns=column_count,
            non_empty_cells=non_empty_cells,
            formula_cells=formula_cells,
            preview_rows=self._spreadsheet_preview_rows(values_worksheet),
        )

    def _count_non_empty_cells(self, worksheet: _WorksheetLike) -> int:
        count = 0
        for row in worksheet.iter_rows(values_only=True):
            count += sum(
                1 for value in row if value is not None and str(value).strip() != ""
            )
        return count

    def _count_formula_cells(self, worksheet: _WorksheetLike) -> int:
        count = 0
        for row in worksheet.iter_rows(values_only=True):
            count += sum(
                1 for value in row if isinstance(value, str) and value.startswith("=")
            )
        return count

    def _spreadsheet_preview_rows(self, worksheet: _WorksheetLike) -> list[str]:
        rows: list[str] = []
        max_row = min(int(worksheet.max_row or 0), 20)
        max_col = min(int(worksheet.max_column or 0), 10)
        for index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ),
            start=1,
        ):
            values = [
                self._format_spreadsheet_cell(value)
                for value in row
                if value is not None and str(value).strip() != ""
            ]
            if values:
                rows.append(f"R{index}: " + " | ".join(values))
            if len(rows) >= 8:
                break
        return rows

    def _format_spreadsheet_cell(self, value: object) -> str:
        if value is None:
            return ""
        return str(value)

    def _profile_image_collection(
        self,
        context: _ProfileContext,
    ) -> tuple[ImageCollectionProfile | None, list[str]]:
        image_items = [
            item for item in context.files if item.path.suffix.lower() in IMAGE_SUFFIXES
        ]
        if not image_items:
            return None, []

        root = self._common_image_root(image_items)
        format_counts: Counter[str] = Counter()
        resolution_counts: Counter[str] = Counter()
        directory_counts: Counter[str] = Counter()
        warnings: list[str] = []

        for item in image_items:
            suffix = item.path.suffix.lower()
            format_counts[suffix] += 1
            parent = item.path.parent
            if self._is_relative_to(parent.resolve(), root.resolve()):
                directory = parent.relative_to(root).as_posix()
            else:
                directory = parent.name
            directory_counts[directory if directory != "." else root.name] += 1
            try:
                with Image.open(item.path) as image:
                    width, height = image.size
            except (OSError, UnidentifiedImageError) as err:
                warnings.append(f"Failed to profile image {item.relative_path}: {err}")
                continue
            resolution_counts[f"{width}x{height}"] += 1

        return (
            ImageCollectionProfile(
                root=root,
                relative_root=root.name,
                total_images=len(image_items),
                format_counts=dict(format_counts),
                resolution_counts=dict(resolution_counts),
                directory_counts=dict(directory_counts),
                warnings=warnings,
            ),
            warnings,
        )

    def _common_image_root(self, image_items: list[_DiscoveredFile]) -> Path:
        parents = [item.path.parent.resolve() for item in image_items]
        return Path(commonpath([str(path) for path in parents]))

    def _table_role(self, label: str) -> TableRole:
        normalized = Path(label.split("::", 1)[0]).stem.lower()
        tokens = normalized.replace("-", "_").split("_")
        token_set = set(tokens)
        if {"sample", "submission"}.issubset(token_set):
            return TableRole.SAMPLE_SUBMISSION
        if "train" in token_set or normalized == "training":
            return TableRole.TRAIN
        if "valid" in token_set or "validation" in token_set or "val" in token_set:
            return TableRole.VALIDATION
        if "test" in token_set:
            return TableRole.TEST
        return TableRole.TABLE

    def _profile_table(
        self,
        item: _DiscoveredFile,
        *,
        frame: pd.DataFrame,
        relative_path: str,
        source_label: str,
        role: TableRole,
    ) -> TableProfile:
        table_warnings: list[str] = []

        numeric_summaries = self._numeric_summaries(frame)
        likely_time_columns = self._likely_time_columns(frame)
        likely_categorical_columns = self._likely_categorical_columns(frame)

        return TableProfile(
            path=item.path,
            relative_path=relative_path,
            source_file=item.path,
            source_label=source_label,
            role=role,
            rows=len(frame),
            columns=len(frame.columns),
            column_names=[str(column) for column in frame.columns],
            dtypes={str(column): str(dtype) for column, dtype in frame.dtypes.items()},
            missing_rates={
                str(column): float(rate)
                for column, rate in frame.isna().mean(numeric_only=False).items()
            },
            unique_counts=self._unique_counts(frame),
            numeric_summaries=numeric_summaries,
            likely_time_columns=likely_time_columns,
            likely_categorical_columns=likely_categorical_columns,
            sample_rows=self._sample_rows(frame),
            warnings=table_warnings,
        )

    def _read_table(self, path: Path) -> pd.DataFrame:
        suffix = path.suffix.lower()
        match suffix:
            case ".csv":
                return pd.read_csv(
                    path,
                    index_col=0
                    if self._first_header_cell_is_blank(path, ",")
                    else None,
                )
            case ".tsv":
                return pd.read_csv(
                    path,
                    sep="\t",
                    index_col=0
                    if self._first_header_cell_is_blank(path, "\t")
                    else None,
                )
            case ".xlsx" | ".xls":
                return pd.read_excel(path)
            case ".parquet":
                return pd.read_parquet(path)

        msg = f"Unsupported tabular suffix: {suffix}"
        raise ValueError(msg)

    def _first_header_cell_is_blank(self, path: Path, delimiter: str) -> bool:
        with path.open(newline="") as f:
            first_row = next(csv.reader(f, delimiter=delimiter), [])
        return bool(first_row) and first_row[0].strip() == ""

    def _numeric_summaries(self, frame: pd.DataFrame) -> dict[str, dict[str, float]]:
        summaries: dict[str, dict[str, float]] = {}
        for column in frame.columns:
            series = frame[column]
            if not is_numeric_dtype(series):
                continue
            stats = series.describe()
            summary: dict[str, float] = {}
            for key in ("mean", "std", "min", "25%", "50%", "75%", "max"):
                value = stats.get(key)
                if pd.notna(value):
                    summary[key] = float(value)
            summaries[str(column)] = summary
        return summaries

    def _likely_time_columns(self, frame: pd.DataFrame) -> list[str]:
        likely: list[str] = []
        time_markers = ("date", "time", "timestamp", "datetime")
        for column in frame.columns:
            column_name = str(column)
            series = frame[column]
            lower_name = column_name.lower()
            if is_datetime64_any_dtype(series):
                likely.append(column_name)
                continue
            if any(marker in lower_name for marker in time_markers):
                parsed = pd.to_datetime(series.dropna().head(20), errors="coerce")
                if not parsed.empty and parsed.notna().mean() >= 0.8:
                    likely.append(column_name)
        return likely

    def _likely_categorical_columns(self, frame: pd.DataFrame) -> list[str]:
        likely: list[str] = []
        row_count = max(len(frame), 1)
        for column in frame.columns:
            series = frame[column]
            try:
                unique_count = int(series.nunique(dropna=True))
            except TypeError as err:
                msg = f"categorical detection failed for {column}: {err}"
                raise _TableProfilingError(msg) from err
            if not is_numeric_dtype(series) and unique_count <= max(20, row_count // 2):
                likely.append(str(column))
        return likely

    def _unique_counts(self, frame: pd.DataFrame) -> dict[str, int]:
        try:
            counts = frame.nunique(dropna=True)
        except TypeError as err:
            msg = f"unique counts failed: {err}"
            raise _TableProfilingError(msg) from err
        return {str(column): int(count) for column, count in counts.items()}

    def _sample_rows(self, frame: pd.DataFrame) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for raw_row in frame.head(self.sample_rows).to_dict(orient="records"):
            row: dict[str, object] = {}
            for key, value in raw_row.items():
                row[str(key)] = None if pd.isna(value) else value
            rows.append(row)
        return rows

    def _is_relative_to(self, path: Path, parent: Path) -> bool:
        try:
            path.relative_to(parent)
        except ValueError:
            return False
        return True
